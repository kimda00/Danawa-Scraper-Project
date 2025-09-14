import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO
from tqdm import tqdm

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
}

# 검색어
search_keyword = "후드티"

# 다나와 검색 결과 URL
base_url = f"https://search.danawa.com/dsearch.php?query={search_keyword}"

# HTTP 요청 보내기
response = requests.get(base_url, headers=headers)
response.raise_for_status()

# HTML 파싱
soup = BeautifulSoup(response.text, 'html.parser')

# 상품명, 가격, 링크, 썸네일 정보 추출
product_elements = soup.select("a.click_log_product_standard_title_")
product_prices = soup.select("a.click_log_product_standard_price_ > strong")
product_images = soup.select("a.thumb_link img")

# 엑셀 파일 생성
wb = Workbook()
main_sheet = wb.active
main_sheet.title = '검색 결과'

# 헤더 작성
main_sheet.append(['상품명', '가격', '썸네일', '링크', '상세 정보'])

# 검색 결과 저장
len_product = len(product_elements)
for i, (product, price_element, img) in enumerate(tqdm(zip(product_elements, product_prices, product_images), total=len_product), start=2):
    product_name = product.text.strip()
    product_price = price_element.text.strip()
    product_link = product['href']
    if img.get('class') and "lazyload" in img.get('class'):
      product_thumbnail = img.get('data-src')
    else:
      product_thumbnail = img.get('src')

    main_sheet[f'A{i}'] = product_name
    main_sheet[f'B{i}'] = product_price
    main_sheet[f'D{i}'] = product_link

    detail_sheet_name = f"{product_name[:15]}"
    # 링크 연결
    main_sheet[f'E{i}'] = f"=HYPERLINK(\"#'{detail_sheet_name}'!A1\", \"상세 정보 보기\")"

    # 상세 페이지 접근
    try:
        detail_response = requests.get(product_link, headers=headers)
        detail_response.raise_for_status()
        detail_soup = BeautifulSoup(detail_response.text, 'html.parser')

        # 판매 업체 정보 추출
        seller_elements = detail_soup.select("div.box__logo > img")
        price_elements = detail_soup.select("div.sell-price > span.text__num")
        delivery_elements = detail_soup.select("div.box__delivery")

        # 상세 정보 저장
        detail_sheet = wb.create_sheet(title=detail_sheet_name)
        detail_sheet.append(['판매 업체', '가격', '배달비'])

        for seller, price, delivery in zip(seller_elements, price_elements, delivery_elements):
            seller_name = seller.get('alt')
            seller_price = price.text.strip()
            seller_delivery = delivery.text.strip()
            seller_delivery = seller_delivery.replace('원', '')
            detail_sheet.append([seller_name, seller_price, seller_delivery])

    except Exception as e:
        print(f"상세 정보 처리 중 오류 발생: {e}")

    # 이미지 저장 및 서식 지정
    if product_thumbnail:
        try:
            img_response = requests.get(f"https:{product_thumbnail}")
            img_response.raise_for_status()
            img_data = Image(BytesIO(img_response.content))
            img_data.width, img_data.height = 140, 140
            main_sheet.add_image(img_data, f'C{i}')
            main_sheet.row_dimensions[i].height = 140
        except Exception as e:
            print(f"Error: {e}")

# 엑셀 스타일링
main_sheet.column_dimensions['A'].width = 30
main_sheet.column_dimensions['B'].width = 15
main_sheet.column_dimensions['C'].width = 20
main_sheet.column_dimensions['D'].width = 50
main_sheet.column_dimensions['E'].width = 20

# 저장
wb.save("검색_및_상세_결과.xlsx")
print("엑셀 파일 저장 완료.")